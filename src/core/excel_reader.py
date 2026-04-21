"""Excel文件读取器

支持 .xlsx (openpyxl) 和 .xls (xlrd) 格式，支持多sheet读取
"""

import os
from datetime import datetime, timedelta
from typing import List, Any, Optional, Tuple, Dict

from .exceptions import ExcelFormatError


class SheetData:
    """单个sheet的数据封装"""

    def __init__(self, name: str, data: List[List[Any]]):
        self.name = name
        self.data = data

    def has_data(self) -> bool:
        """检查sheet是否有有效数据（至少2行：表头+数据）"""
        return len(self.data) >= 2

    def __len__(self) -> int:
        return len(self.data)


class ExcelReader:
    """读取Excel文件并返回结构化数据，支持多sheet"""

    def __init__(self):
        self.file_path: Optional[str] = None
        self.sheets: List[SheetData] = []
        self.sheet_name: str = ""  # 保持向后兼容

    def read(self, file_path: str) -> Tuple[List[List[Any]], str]:
        """读取Excel文件并返回第一个sheet的数据（向后兼容）

        Args:
            file_path: Excel文件路径

        Returns:
            (数据列表, sheet名称)

        Raises:
            ExcelFormatError: 文件不存在或格式错误
        """
        sheets = self.read_all_sheets(file_path)
        if not sheets:
            raise ExcelFormatError(
                'Excel文件没有包含有效数据的sheet',
                error_code='EMPTY_FILE',
                suggestion='请确保文件至少有一个sheet包含数据行（不仅是空表头）。'
            )
        # 返回第一个sheet的数据（向后兼容）
        first_sheet = sheets[0]
        self.sheet_name = first_sheet.name
        return first_sheet.data, first_sheet.name

    def read_all_sheets(self, file_path: str) -> List[SheetData]:
        """读取Excel文件所有sheet，过滤空sheet

        Args:
            file_path: Excel文件路径

        Returns:
            SheetData列表（已过滤空sheet）

        Raises:
            ExcelFormatError: 文件不存在或格式错误
        """
        if not os.path.exists(file_path):
            raise ExcelFormatError(
                f'文件不存在：{file_path}',
                error_code='FILE_NOT_FOUND',
                suggestion='请确认文件路径是否正确，或检查文件是否被移动/删除。'
            )

        self.file_path = file_path

        # 根据扩展名选择读取方法
        if file_path.lower().endswith('.xlsx'):
            sheets = self._read_xlsx_all_sheets(file_path)
        elif file_path.lower().endswith('.xls'):
            sheets = self._read_xls_all_sheets(file_path)
        else:
            ext = os.path.splitext(file_path)[1] or '未知'
            raise ExcelFormatError(
                f'不支持的文件格式"{ext}"',
                error_code='UNSUPPORTED_FORMAT',
                suggestion='请使用.xlsx或.xls格式的Excel文件。'
            )

        # 过滤掉没有数据的sheet（至少2行：表头+数据）
        self.sheets = [s for s in sheets if s.has_data()]
        return self.sheets

    def _read_xlsx_all_sheets(self, file_path: str) -> List[SheetData]:
        """使用 openpyxl 读取 .xlsx 文件的所有sheet"""
        try:
            import openpyxl
        except ImportError:
            raise ExcelFormatError(
                '缺少必要的依赖库：openpyxl',
                error_code='DEPENDENCY_MISSING',
                suggestion='请运行命令安装：pip install openpyxl>=3.1.2'
            )

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                data = []

                for row in sheet.iter_rows(values_only=True):
                    row_data = []
                    for cell in row:
                        if cell is None:
                            row_data.append("")
                        elif isinstance(cell, (int, float)):
                            # Check if it looks like an Excel date serial
                            if 40000 <= cell <= 60000:
                                try:
                                    # Convert Excel serial to date
                                    # Excel epoch is 1899-12-30 (with 1900 leap year bug)
                                    excel_epoch = datetime(1899, 12, 30)
                                    dt = excel_epoch + timedelta(days=float(cell))
                                    row_data.append(dt.strftime('%Y-%m-%d'))
                                except:
                                    row_data.append(str(cell))
                            else:
                                row_data.append(str(cell))
                        elif isinstance(cell, datetime):
                            # Already a datetime object
                            row_data.append(cell.strftime('%Y-%m-%d'))
                        else:
                            row_data.append(str(cell))
                    data.append(row_data)

                sheets.append(SheetData(sheet_name, data))

            return sheets

        except Exception as e:
            # 检查是否为密码保护
            error_str = str(e).lower()
            if 'password' in error_str or 'protected' in error_str:
                raise ExcelFormatError(
                    '无法打开受密码保护的Excel文件',
                    error_code='PASSWORD_PROTECTED',
                    suggestion='请先移除密码保护，或用Excel打开后另存为新文件（不带密码）。'
                )
            # 其他解析错误
            raise ExcelFormatError(
                '无法读取Excel文件，文件可能已损坏',
                error_code='CORRUPTED_FILE',
                suggestion='请尝试用Excel打开该文件，然后另存为新文件后再试。'
            )

    def _read_xls_all_sheets(self, file_path: str) -> List[SheetData]:
        """使用 xlrd 读取 .xls 文件的所有sheet"""
        try:
            import xlrd
        except ImportError:
            raise ExcelFormatError(
                '缺少必要的依赖库：xlrd',
                error_code='DEPENDENCY_MISSING',
                suggestion='请运行命令安装：pip install xlrd>=2.0.1'
            )

        try:
            workbook = xlrd.open_workbook(file_path)
            sheets = []

            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                data = []

                for row_idx in range(sheet.nrows):
                    row = sheet.row_values(row_idx)
                    # 将每个单元格转换为字符串
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    data.append(row_data)

                sheets.append(SheetData(sheet.name, data))

            return sheets

        except Exception as e:
            # 检查是否为密码保护
            error_str = str(e).lower()
            if 'password' in error_str or 'protected' in error_str:
                raise ExcelFormatError(
                    '无法打开受密码保护的Excel文件',
                    error_code='PASSWORD_PROTECTED',
                    suggestion='请先移除密码保护，或用Excel打开后另存为新文件（不带密码）。'
                )
            # 其他解析错误
            raise ExcelFormatError(
                '无法读取Excel文件，文件可能已损坏',
                error_code='CORRUPTED_FILE',
                suggestion='请尝试用Excel打开该文件，然后另存为新文件后再试。'
            )
