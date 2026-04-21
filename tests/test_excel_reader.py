"""ExcelReader 单元测试"""

import pytest
import os
import tempfile
from src.core.excel_reader import ExcelReader


class TestExcelReader:
    """测试 ExcelReader 类"""

    @pytest.fixture
    def reader(self):
        return ExcelReader()

    def test_read_xlsx(self, reader, tmp_path):
        """测试读取 .xlsx 文件"""
        # 创建临时 xlsx 文件
        try:
            import openpyxl
        except ImportError:
            pytest.skip("未安装 openpyxl")

        xlsx_file = tmp_path / "test.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['日期', '发货人'])
        sheet.append(['2024-01-15', '张三'])
        workbook.save(xlsx_file)

        # 读取 - 返回 (data, sheet_name)
        data, sheet_name = reader.read(str(xlsx_file))

        # 验证
        assert len(data) == 2
        assert data[0] == ['日期', '发货人']
        assert data[1] == ['2024-01-15', '张三']
        assert sheet_name == 'Sheet'

    def test_read_xls(self, reader, tmp_path):
        """测试读取 .xls 文件"""
        # 创建临时 xls 文件
        try:
            import xlwt
        except ImportError:
            pytest.skip("未安装 xlwt")

        xls_file = tmp_path / "test.xls"
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('test')
        sheet.write(0, 0, '日期')
        sheet.write(0, 1, '发货人')
        sheet.write(1, 0, '2024-01-15')
        sheet.write(1, 1, '张三')
        workbook.save(xls_file)

        # 读取 - 返回 (data, sheet_name)
        data, sheet_name = reader.read(str(xls_file))

        # 验证
        assert len(data) == 2
        assert data[0] == ['日期', '发货人']
        assert data[1] == ['2024-01-15', '张三']
        assert sheet_name == 'test'

    def test_file_not_found(self, reader):
        """测试文件不存在时的错误处理"""
        from src.core.exceptions import ExcelFormatError
        with pytest.raises(ExcelFormatError) as exc_info:
            reader.read("/不存在的路径/文件.xlsx")
        
        # 验证中文错误消息
        assert "文件不存在" in str(exc_info.value)

    def test_invalid_extension(self, reader, tmp_path):
        """测试不支持的文件格式"""
        from src.core.exceptions import ExcelFormatError
        txt_file = tmp_path / "test.txt"
        txt_file.write_text("这不是Excel文件")

        with pytest.raises(ExcelFormatError) as exc_info:
            reader.read(str(txt_file))
        
        # 验证中文错误消息
        assert "不支持的文件格式" in str(exc_info.value)

    def test_empty_file_handling(self, reader, tmp_path):
        """测试空单元格处理"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("未安装 openpyxl")

        xlsx_file = tmp_path / "empty.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['日期', '发货人'])
        sheet.append(['2024-01-15', None])  # 空单元格
        workbook.save(xlsx_file)

        data, sheet_name = reader.read(str(xlsx_file))

        # 空单元格应转为空字符串
        assert data[1][1] == ""

    def test_read_all_sheets(self, reader, tmp_path):
        """测试读取多个sheet"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("未安装 openpyxl")

        xlsx_file = tmp_path / "multi_sheet.xlsx"
        workbook = openpyxl.Workbook()
        
        # Sheet1
        sheet1 = workbook.active
        sheet1.title = "Sheet1"
        sheet1.append(['日期', '发货人'])
        sheet1.append(['2024-01-15', '张三'])
        
        # Sheet2
        sheet2 = workbook.create_sheet("Sheet2")
        sheet2.append(['日期', '发货人'])
        sheet2.append(['2024-01-16', '李四'])
        
        # 空Sheet3（应该被忽略）
        sheet3 = workbook.create_sheet("Sheet3")
        
        workbook.save(xlsx_file)

        # 读取所有sheet
        sheets = reader.read_all_sheets(str(xlsx_file))
        
        # 验证
        assert len(sheets) == 2  # 只有2个有效sheet
        assert sheets[0].name == 'Sheet1'
        assert len(sheets[0].data) == 2
        assert sheets[1].name == 'Sheet2'
        assert len(sheets[1].data) == 2
